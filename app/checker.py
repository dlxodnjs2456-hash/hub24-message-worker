import asyncio, csv, io, math, re, time
from datetime import datetime, timezone
from pathlib import Path
import httpx
from fastapi import UploadFile, File, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook

from .main import app, auth
from . import db
from .settings import settings

_checker_tasks={}

def now_iso(): return datetime.now(timezone.utc).isoformat()
def q(t): return db.sb.table(t)

def _display_phone(v:str):
    d=re.sub(r'\D','',v or '')
    if d.startswith('82') and len(d)>=12: d='0'+d[2:]
    if len(d)==11 and d.startswith('010'): return f'{d[:3]}-{d[3:7]}-{d[7:]}'
    return v

def _normalize(v):
    d=re.sub(r'\D','',str(v or ''))
    if d.startswith('82'): d='0'+d[2:]
    if len(d)==11 and d.startswith('010'): return '+82'+d[1:]
    return None

def _provider_phone(v):
    n=_normalize(v)
    return n[1:] if n and n.startswith('+') else n

def _read_numbers(name,raw):
    ext=Path(name or '').suffix.lower(); out=[]
    if ext in ('.xlsx','.xlsm'):
        wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True);ws=wb.active
        first=next(ws.iter_rows(values_only=True),None)
        if first:
            headers=[str(x or '').strip().lower() for x in first]
            aliases={'전화번호','휴대폰','휴대전화','연락처','phone','phone_number','mobile','mobile_number','tel'}
            idx=next((i for i,h in enumerate(headers) if h in aliases),None)
            if idx is not None:
                for rn,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
                    if idx<len(row) and row[idx] is not None: out.append((rn,str(row[idx])))
            else:
                for rn,row in enumerate(ws.iter_rows(values_only=True),start=1):
                    for v in row:
                        if v is not None and len(re.sub(r'\D','',str(v)))>=10: out.append((rn,str(v)));break
    else:
        text=raw.decode('utf-8-sig',errors='ignore')
        rows=list(csv.reader(io.StringIO(text))) if ext=='.csv' else [[x] for x in text.splitlines()]
        for rn,row in enumerate(rows,start=1):
            for v in row:
                if len(re.sub(r'\D','',str(v)))>=10: out.append((rn,str(v)));break
    return out

def _rate(count):
    if count<1000:return 0
    if count<8000:return 1.0
    if count<50000:return .8
    if count<100000:return .7
    return .6

def _api_ready(): return bool(settings.check_api_base_url and settings.check_api_key)
def _url(path): return settings.check_api_base_url.rstrip('/')+'/'+path.lstrip('/')
def _headers(): return {'token':settings.check_api_key}

def _task_id(data):
    if isinstance(data,dict):
        for k in ('task_id','taskId','id'):
            if data.get(k) not in (None,''): return str(data[k])
        d=data.get('data')
        if isinstance(d,dict):
            for k in ('task_id','taskId','id'):
                if d.get(k) not in (None,''): return str(d[k])
        if isinstance(d,(str,int)) and str(d).strip(): return str(d)
    return None

def _status_value(data):
    if not isinstance(data,dict): return ''
    d=data.get('data') if isinstance(data.get('data'),dict) else data
    for k in ('status','state','task_status','taskStatus'):
        if k in d and d.get(k) is not None:return str(d.get(k)).strip().lower()
    return ''

def _is_done(data):
    if not isinstance(data,dict): return False
    d=data.get('data') if isinstance(data.get('data'),dict) else data
    if d.get('finish') is True or d.get('finished') is True or d.get('completed') is True:return True
    s=_status_value(data)
    if s in {'2','3','done','finish','finished','complete','completed','success','successful'}:return True
    total=d.get('total') or d.get('total_count') or d.get('count')
    done=d.get('done') or d.get('completed_count') or d.get('finish_count')
    try:
        if total is not None and done is not None and int(total)>0 and int(done)>=int(total):return True
    except Exception:pass
    return False

def _rows_from_export(raw:bytes):
    rows=[]
    if raw[:2]==b'PK':
        try:
            wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True);ws=wb.active
            rows=[list(r) for r in ws.iter_rows(values_only=True)]
            return rows
        except Exception:pass
    text=None
    for enc in ('utf-8-sig','utf-8','gb18030','cp949'):
        try:text=raw.decode(enc);break
        except Exception:continue
    if text is None:return rows
    sample=text[:4096]
    delim='\t' if sample.count('\t')>sample.count(',') else ','
    try: rows=[r for r in csv.reader(io.StringIO(text),delimiter=delim) if any(str(x or '').strip() for x in r)]
    except Exception: rows=[[x] for x in text.splitlines() if x.strip()]
    return rows

def _parse_export(raw:bytes):
    rows=_rows_from_export(raw)
    if not rows:return {}
    aliases_phone={'phone','phone_number','mobile','number','tel','telephone','手机号','手机','电话号码','电话','전화번호'}
    aliases_uid={'telegram_id','telegram_user_id','uid','user_id','tg_id','id'}
    aliases_user={'telegram_username','username','tg_username','user_name'}
    aliases_active={'telegram_active','last_seen','active_at','last_online','online_time','active'}
    first=[str(x or '').strip().lower() for x in rows[0]]
    has_header=any(x in aliases_phone|aliases_uid|aliases_user|aliases_active for x in first)
    start=1 if has_header else 0
    def idx(aliases,default=None):
        if has_header:
            for i,h in enumerate(first):
                if h in aliases:return i
        return default
    pi=idx(aliases_phone,0);ui=idx(aliases_uid,1 if not has_header else None);ni=idx(aliases_user,2 if not has_header else None);ai=idx(aliases_active,3 if not has_header else None)
    out={}
    for row in rows[start:]:
        if pi is None or pi>=len(row):continue
        n=_normalize(str(row[pi] or ''))
        if not n:continue
        out[n]={'status':'REGISTERED','telegram_id':row[ui] if ui is not None and ui<len(row) else None,'telegram_username':row[ni] if ni is not None and ni<len(row) else None,'telegram_active':str(row[ai]) if ai is not None and ai<len(row) and row[ai] is not None else None}
    return out

async def _submit_provider_task(client,phones,seq):
    body='\n'.join((_provider_phone(x) or '') for x in phones if _provider_phone(x))+'\n'
    files={'file':(f'npay_checker_{seq}.txt',body.encode('utf-8'),'text/plain')}
    data={'filter_type':str(settings.check_api_filter_type or '1'),'conuntry':str(settings.check_api_country or 'KR')}
    r=await client.post(_url('/addTask'),headers=_headers(),data=data,files=files)
    if r.status_code>=400:raise RuntimeError(f'ADD_TASK_HTTP_{r.status_code}')
    try:j=r.json()
    except Exception as e:raise RuntimeError('ADD_TASK_INVALID_JSON') from e
    tid=_task_id(j)
    if not tid:raise RuntimeError(f'ADD_TASK_NO_TASK_ID:{str(j)[:300]}')
    return tid

async def _wait_provider_task(client,tid):
    interval=max(1,int(settings.check_api_poll_interval_seconds or 5));max_wait=max(interval,int(settings.check_api_max_wait_seconds or 900));started=time.monotonic();last=''
    while time.monotonic()-started<max_wait:
        r=await client.get(_url('/checkTask'),headers=_headers(),params={'id':tid})
        if r.status_code==429:await asyncio.sleep(interval);continue
        if r.status_code>=400:raise RuntimeError(f'CHECK_TASK_HTTP_{r.status_code}')
        try:j=r.json()
        except Exception as e:raise RuntimeError('CHECK_TASK_INVALID_JSON') from e
        last=_status_value(j) or last
        if _is_done(j):return last or 'COMPLETED'
        await asyncio.sleep(interval)
    raise TimeoutError(f'CHECK_TASK_TIMEOUT:{tid}')

async def _export_provider_task(client,tid):
    r=await client.get(_url('/exportPhone'),headers=_headers(),params={'id':tid})
    if r.status_code>=400:raise RuntimeError(f'EXPORT_PHONE_HTTP_{r.status_code}')
    ctype=(r.headers.get('content-type') or '').lower()
    if 'json' in ctype:
        try:j=r.json()
        except Exception:j={}
        d=j.get('data') if isinstance(j,dict) else None
        cand=[]
        if isinstance(j,dict):cand += [j.get('url'),j.get('download_url'),j.get('file_url')]
        if isinstance(d,dict):cand += [d.get('url'),d.get('download_url'),d.get('file_url')]
        url=next((x for x in cand if isinstance(x,str) and x.startswith('http')),None)
        if url:
            rr=await client.get(url,headers=_headers());rr.raise_for_status();return rr.content
    return r.content

async def _process(job_id,user):
    q('npay_checker_jobs').update({'status':'RUNNING','started_at':now_iso(),'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
    timeout=max(5,int(settings.check_api_timeout_seconds or 30));batch_size=max(1,min(int(settings.check_api_batch_size or 5000),100000));task_ids=[]
    try:
        items=q('npay_checker_items').select('id,normalized_phone').eq('job_id',job_id).eq('user_id',user).order('id').execute().data or []
        by_phone={x['normalized_phone']:x for x in items};found={}
        async with httpx.AsyncClient(timeout=timeout,follow_redirects=True) as client:
            for seq,start in enumerate(range(0,len(items),batch_size),start=1):
                chunk=items[start:start+batch_size];phones=[x['normalized_phone'] for x in chunk]
                tid=await _submit_provider_task(client,phones,seq);task_ids.append(tid)
                q('npay_checker_jobs').update({'api_task_ids':task_ids,'provider_status':'SUBMITTED','provider_last_check_at':now_iso(),'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
                st=await _wait_provider_task(client,tid)
                q('npay_checker_jobs').update({'provider_status':st,'provider_last_check_at':now_iso(),'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
                export=await _export_provider_task(client,tid);found.update(_parse_export(export))
        checked=now_iso()
        for phone,x in by_phone.items():
            v=found.get(phone)
            upd={'checked_at':checked,'api_status':'200'}
            if v:upd.update(v)
            else:upd.update({'status':'NOT_REGISTERED','telegram_id':None,'telegram_username':None,'telegram_active':None})
            q('npay_checker_items').update(upd).eq('id',x['id']).eq('user_id',user).execute()
        rows=q('npay_checker_items').select('status').eq('job_id',job_id).eq('user_id',user).execute().data or []
        done=len(rows);reg=sum(1 for x in rows if x['status']=='REGISTERED');unk=sum(1 for x in rows if x['status'] in ('UNKNOWN','NOT_REGISTERED'));err=sum(1 for x in rows if x['status'] in ('API_ERROR','RATE_LIMITED','TIMEOUT'))
        q('npay_checker_jobs').update({'status':'COMPLETED','provider_status':'COMPLETED','completed_count':done,'registered_count':reg,'unknown_count':unk,'error_count':err,'completed_at':now_iso(),'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
    except TimeoutError as e:
        q('npay_checker_items').update({'status':'TIMEOUT','api_status':'TIMEOUT','checked_at':now_iso(),'error_message':'검수 API 최대 대기시간 초과'}).eq('job_id',job_id).eq('user_id',user).eq('status','WAITING').execute()
        q('npay_checker_jobs').update({'status':'FAILED','provider_status':'TIMEOUT','error_message':str(e)[:1000],'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
    except Exception as e:
        q('npay_checker_items').update({'status':'API_ERROR','api_status':'ERROR','checked_at':now_iso(),'error_message':str(e)[:500]}).eq('job_id',job_id).eq('user_id',user).eq('status','WAITING').execute()
        q('npay_checker_jobs').update({'status':'FAILED','provider_status':'ERROR','error_message':str(e)[:1000],'updated_at':now_iso()}).eq('id',job_id).eq('user_id',user).execute()
    finally:_checker_tasks.pop(str(job_id),None)

@app.get('/v1/checker/config')
def checker_config(user=Depends(auth)):
    return {'minimum_count':1000,'maximum_count':1000000,'tiers':[{'min':1000,'max':7999,'unit_price':1.0},{'min':8000,'max':49999,'unit_price':.8},{'min':50000,'max':99999,'unit_price':.7},{'min':100000,'max':None,'unit_price':.6}],'api_ready':_api_ready(),'provider_mode':'TASK_POLL_EXPORT','country':settings.check_api_country,'filter_type':settings.check_api_filter_type}

@app.post('/v1/checker/upload')
async def checker_upload(file:UploadFile=File(...),user=Depends(auth)):
    raw=await file.read()
    if len(raw)>25*1024*1024: raise HTTPException(413,'FILE_TOO_LARGE')
    vals=_read_numbers(file.filename or 'phones.xlsx',raw);seen=set();items=[];invalid=0;dup=0
    for rn,v in vals:
        n=_normalize(v)
        if not n: invalid+=1;continue
        if n in seen: dup+=1;continue
        seen.add(n);items.append((rn,v,n))
    count=len(items)
    if count<1000: raise HTTPException(400,f'MINIMUM_CHECK_COUNT_1000: valid={count}')
    if count>1000000: raise HTTPException(400,'MAXIMUM_CHECK_COUNT_1000000')
    rate=_rate(count);est=count*rate
    job=q('npay_checker_jobs').insert({'user_id':user,'original_filename':file.filename,'status':'DRAFT','uploaded_count':len(vals),'invalid_count':invalid,'duplicate_count':dup,'requested_count':count,'unit_price':rate,'estimated_cost':est}).execute().data[0]
    batch=[]
    for rn,v,n in items:
        batch.append({'job_id':job['id'],'user_id':user,'row_no':rn,'phone':_display_phone(v),'normalized_phone':n,'status':'WAITING'})
        if len(batch)>=1000:q('npay_checker_items').insert(batch).execute();batch=[]
    if batch:q('npay_checker_items').insert(batch).execute()
    return {'job':job,'quote':{'requested_count':count,'unit_price':rate,'estimated_cost':est,'charged_points':math.ceil(est),'invalid_count':invalid,'duplicate_count':dup}}

@app.get('/v1/checker/jobs')
def checker_jobs(user=Depends(auth)):return {'items':q('npay_checker_jobs').select('*').eq('user_id',user).order('created_at',desc=True).limit(100).execute().data or []}

@app.get('/v1/checker/jobs/{jid}')
def checker_job(jid:int,user=Depends(auth)):
    row=q('npay_checker_jobs').select('*').eq('id',jid).eq('user_id',user).maybe_single().execute().data
    if not row: raise HTTPException(404,'CHECKER_JOB_NOT_FOUND')
    return row

@app.post('/v1/checker/jobs/{jid}/start')
async def checker_start(jid:int,user=Depends(auth)):
    row=q('npay_checker_jobs').select('*').eq('id',jid).eq('user_id',user).maybe_single().execute().data
    if not row: raise HTTPException(404,'CHECKER_JOB_NOT_FOUND')
    if not _api_ready(): raise HTTPException(503,'CHECK_API_NOT_CONFIGURED')
    if row['status']=='DRAFT':
        try:db.sb.rpc('npay_charge_checker_job',{'p_user':user,'p_job_id':jid}).execute()
        except Exception as e: raise HTTPException(400,str(e))
    elif row['status'] not in ('QUEUED','FAILED'): raise HTTPException(409,'CHECKER_JOB_ALREADY_STARTED')
    key=str(jid)
    if key not in _checker_tasks or _checker_tasks[key].done(): _checker_tasks[key]=asyncio.create_task(_process(jid,user))
    return {'ok':True,'job_id':jid,'status':'RUNNING'}

@app.get('/v1/checker/jobs/{jid}/results')
def checker_results(jid:int,limit:int=200,user=Depends(auth)):
    return {'items':q('npay_checker_items').select('id,row_no,phone,status,telegram_id,telegram_username,telegram_active,checked_at,error_code,error_message').eq('job_id',jid).eq('user_id',user).order('id').limit(min(limit,1000)).execute().data or []}

@app.get('/v1/checker/jobs/{jid}/download')
def checker_download(jid:int,filter:str='all',user=Depends(auth)):
    job=q('npay_checker_jobs').select('*').eq('id',jid).eq('user_id',user).maybe_single().execute().data
    if not job: raise HTTPException(404,'CHECKER_JOB_NOT_FOUND')
    rows=q('npay_checker_items').select('*').eq('job_id',jid).eq('user_id',user).order('telegram_active',desc=True,nullsfirst=False).execute().data or []
    if filter=='registered': rows=[x for x in rows if x['status']=='REGISTERED']
    elif filter=='unknown': rows=[x for x in rows if x['status'] in ('UNKNOWN','NOT_REGISTERED')]
    elif filter=='error': rows=[x for x in rows if x['status'] in ('API_ERROR','RATE_LIMITED','TIMEOUT')]
    wb=Workbook();ws=wb.active;ws.title='검수결과';ws.append(['전화번호','가입여부','텔레그램 UID','텔레그램 ID','텔레그램 접속일자'])
    labels={'REGISTERED':'가입 확인','NOT_REGISTERED':'미가입','UNKNOWN':'확인 불가','API_ERROR':'오류','RATE_LIMITED':'호출 제한','TIMEOUT':'시간 초과','WAITING':'대기'}
    for x in rows: ws.append([x['phone'],labels.get(x['status'],x['status']),x.get('telegram_id'),x.get('telegram_username'),x.get('telegram_active')])
    bio=io.BytesIO();wb.save(bio);bio.seek(0);fn=f'npay_checker_{jid}_{filter}.xlsx'
    return StreamingResponse(bio,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename="{fn}"'})
