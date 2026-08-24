import asyncio, io, random, re, uuid
from datetime import datetime, timezone
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, Header, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError
from telethon.tl import functions
from telethon.tl.types import InputPhoneContact
from .settings import settings
from .security import enc, dec
from . import db
from .telegram import session_file, proxy_from_url, norm_phone, client_from_account, prepare_source, is_rate_error

app=FastAPI(title='HUB24 MESSAGE Worker',version='5.0.0')
_origins={x.strip() for x in settings.cors_origins.split(',') if x.strip()}
app.add_middleware(CORSMiddleware,allow_origins=sorted(_origins),allow_credentials=True,allow_methods=['*'],allow_headers=['*'])
_tasks={};_stops={};_pauses={}

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def _verify_supabase_jwt(token:str):
    if not token:raise HTTPException(401,'missing bearer token')
    try:
        r=db.sb.auth.get_user(token)
        if not r or not r.user or not r.user.id:raise HTTPException(401,'invalid session')
        return str(r.user.id)
    except HTTPException:raise
    except Exception:raise HTTPException(401,'invalid session')

def auth(authorization:str=Header(default='')):
    if not authorization.lower().startswith('bearer '):raise HTTPException(401,'missing bearer token')
    return _verify_supabase_jwt(authorization.split(' ',1)[1].strip())

def event(user,job,level,scope,msg,account=None):db.event(user,job,level,scope,msg,account)

def mask_phone(p):
    d=norm_phone(p or '')
    return d[:3]+'-****-'+d[-4:] if len(d)>=7 else '***'

def account_view(a):
    a=dict(a);a['phone_masked']=mask_phone(a.get('phone'));a['proxy_enabled']=bool(a.get('proxy_url_enc'))
    for k in ('api_hash_enc','proxy_url_enc','session_path'):a.pop(k,None)
    return a

class ConnectStart(BaseModel):
    api_id:int;api_hash:str;phone:str;proxy_url:str|None=None;label:str|None=None
class ConnectVerify(BaseModel):
    challenge_id:str;code:str;password:str|None=None
class JobCreate(BaseModel):
    batch_id:int;operation_mode:str='CONTACT_AND_SEND';message_text:str;button_text:str;button_url:str;bot_username:str;bot_token:str;delay_min:float=2;delay_max:float=5;global_dedupe:bool=True

@app.get('/health')
def health():return {'ok':True,'service':'hub24-worker','version':'5.0.0','database':'supabase'}

@app.get('/v1/accounts')
def accounts(user=Depends(auth)):
    return {'items':[account_view(x) for x in db.rows('telegram_accounts',user,order='created_at',desc=True)]}

@app.post('/v1/accounts/connect/start')
async def connect_start(p:ConnectStart,user=Depends(auth)):
    cid=uuid.uuid4().hex;spath=session_file(user,cid)
    c=TelegramClient(spath,p.api_id,p.api_hash,proxy=proxy_from_url(p.proxy_url));await c.connect()
    try:sent=await c.send_code_request(p.phone)
    finally:await c.disconnect()
    db.insert('telegram_connect_challenges',{'id':cid,'user_id':user,'label':p.label or p.phone,'phone':p.phone,'api_id':p.api_id,'api_hash_enc':enc(p.api_hash),'proxy_url_enc':enc(p.proxy_url) if p.proxy_url else None,'session_path':spath,'phone_code_hash':sent.phone_code_hash})
    return {'challenge_id':cid,'code_sent':True}

@app.post('/v1/accounts/connect/verify')
async def connect_verify(p:ConnectVerify,user=Depends(auth)):
    ch=db.one('telegram_connect_challenges',user,eq={'id':p.challenge_id})
    if not ch:raise HTTPException(404,'challenge not found')
    c=TelegramClient(ch['session_path'],ch['api_id'],dec(ch['api_hash_enc']),proxy=proxy_from_url(dec(ch['proxy_url_enc'])) if ch.get('proxy_url_enc') else None);await c.connect()
    try:
        try:await c.sign_in(ch['phone'],p.code,phone_code_hash=ch['phone_code_hash'])
        except SessionPasswordNeededError:
            if not p.password:raise HTTPException(409,'2FA_PASSWORD_REQUIRED')
            await c.sign_in(password=p.password)
        me=await c.get_me()
    finally:await c.disconnect()
    a=db.insert('telegram_accounts',{'user_id':user,'label':ch['label'],'phone':ch['phone'],'api_id':ch['api_id'],'api_hash_enc':ch['api_hash_enc'],'session_path':ch['session_path'],'proxy_url_enc':ch.get('proxy_url_enc'),'status':'READY','telegram_user_id':int(me.id) if me else None})
    db.delete('telegram_connect_challenges',eq={'id':p.challenge_id,'user_id':user})
    return account_view(a)

@app.post('/v1/accounts/{aid}/status')
async def account_status(aid:str,user=Depends(auth)):
    a=db.one('telegram_accounts',user,eq={'id':aid})
    if not a:raise HTTPException(404,'account not found')
    c=await client_from_account(a);status='CHECK_REQUIRED';err=None
    try:await c.connect();me=await c.get_me();status='READY' if me else 'CHECK_REQUIRED'
    except Exception as e:err=str(e)
    finally:
        try:await c.disconnect()
        except Exception:pass
    db.update('telegram_accounts',{'status':status,'last_error':err,'last_check_at':now_iso()},eq={'id':aid,'user_id':user})
    return account_view(db.one('telegram_accounts',user,eq={'id':aid}))

@app.delete('/v1/accounts/{aid}')
def account_delete(aid:str,user=Depends(auth)):
    a=db.one('telegram_accounts',user,eq={'id':aid})
    if not a:raise HTTPException(404,'account not found')
    try:Path(a['session_path']).unlink(missing_ok=True)
    except Exception:pass
    db.delete('telegram_accounts',eq={'id':aid,'user_id':user});return {'ok':True}

def extract_numbers(name,data):
    ext=Path(name).suffix.lower();vals=[]
    if ext in ('.xlsx','.xlsm'):
        wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True);ws=wb.active
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None and len(re.sub(r'\D','',str(v)))>=9:vals.append(str(v));break
    else:
        text=data.decode('utf-8-sig',errors='ignore')
        for line in text.splitlines():
            for v in re.split(r'[,\t; ]+',line.strip()):
                if len(re.sub(r'\D','',v))>=9:vals.append(v);break
    return vals

@app.post('/v1/batches/upload')
async def upload_batch(file:UploadFile=File(...),user=Depends(auth)):
    raw=extract_numbers(file.filename or 'db.txt',await file.read());seen=set();phones=[]
    for v in raw:
        n=norm_phone(v)
        if len(n)>=10 and n not in seen:seen.add(n);phones.append(n)
    batch=db.insert('contact_batches',{'user_id':user,'name':file.filename or 'DB','total_count':len(phones),'duplicate_count':max(0,len(raw)-len(phones))});bid=batch['id']
    db.insert_many('contacts',[{'user_id':user,'batch_id':bid,'phone':p,'normalized_phone':p,'state':'NEW'} for p in phones])
    return {'id':bid,'name':file.filename,'total_count':len(phones),'duplicate_count':max(0,len(raw)-len(phones))}

@app.get('/v1/batches')
def batches(user=Depends(auth)):return {'items':db.rows('contact_batches',user,order='created_at',desc=True)}

@app.post('/v1/jobs')
def create_job(p:JobCreate,user=Depends(auth)):
    batch=db.one('contact_batches',user,eq={'id':p.batch_id})
    if not batch:raise HTTPException(404,'batch not found')
    acc=db.rows('telegram_accounts',user,eq={'status':'READY'},order='created_at')
    if not acc:raise HTTPException(400,'READY SESSION account required')
    job=db.insert('jobs',{'user_id':user,'batch_id':p.batch_id,'status':'WAITING','operation_mode':p.operation_mode,'message_text':p.message_text,'button_text':p.button_text,'button_url':p.button_url,'bot_username':p.bot_username,'bot_token_enc':enc(p.bot_token),'delay_min':p.delay_min,'delay_max':p.delay_max,'global_dedupe':p.global_dedupe,'total_count':int(batch.get('total_count') or 0),'pending_count':int(batch.get('total_count') or 0)})
    jid=job['id'];contacts=db.rows('contacts',user,eq={'batch_id':p.batch_id},order='created_at');targets=[]
    for i,c in enumerate(contacts):targets.append({'user_id':user,'job_id':jid,'contact_id':c['id'],'phone':c['phone'],'telegram_user_id':c.get('telegram_user_id'),'assigned_account_id':acc[i%len(acc)]['id'],'state':'WAITING','stage':'대기'})
    db.insert_many('job_targets',targets);event(user,jid,'INFO','JOB',f'JOB 생성 / 대상 {len(targets)}건 / 계정 {len(acc)}개 고정 배정');return job

@app.get('/v1/jobs')
def jobs(user=Depends(auth)):
    items=db.rows('jobs',user,order='created_at',desc=True);batches={str(b['id']):b['name'] for b in db.rows('contact_batches',user,order=None)}
    for j in items:j['batch_name']=batches.get(str(j.get('batch_id')),j.get('batch_id'))
    return {'items':items}

@app.get('/v1/jobs/{jid}')
def job(jid:str,user=Depends(auth)):
    r=db.one('jobs',user,eq={'id':jid})
    if not r:raise HTTPException(404,'job not found')
    return r

@app.get('/v1/jobs/{jid}/targets')
def targets(jid:str,limit:int=1000,user=Depends(auth)):
    items=db.rows('job_targets',user,eq={'job_id':jid},order='created_at',limit=min(limit,5000));acc={str(a['id']):a for a in db.rows('telegram_accounts',user,order=None)}
    for x in items:
        p=x.get('phone') or '';x['phone_display']=f'{p[:3]}-{p[3:7]}-{p[7:]}' if len(p)==11 else p;x['detail']=x.get('error_detail') or x.get('stage');x['account_label']=(acc.get(str(x.get('assigned_account_id'))) or {}).get('label')
    return {'items':items}

@app.get('/v1/jobs/{jid}/events')
def events(jid:str,limit:int=300,user=Depends(auth)):
    rows=db.rows('job_logs',user,eq={'job_id':jid},order='created_at',desc=True,limit=min(limit,1000));acc={str(a['id']):a for a in db.rows('telegram_accounts',user,order=None)}
    for e in rows:e['account_label']=(acc.get(str(e.get('account_id'))) or {}).get('label')
    return {'items':list(reversed(rows))}

def recalc_job(user,jid):
    ts=db.rows('job_targets',user,eq={'job_id':jid},order=None);sent=sum(1 for x in ts if x.get('state')=='SENT');failed=sum(1 for x in ts if x.get('state')=='FAILED');pending=sum(1 for x in ts if x.get('state') in ('WAITING','PROCESSING'))
    db.update('jobs',{'sent_count':sent,'failed_count':failed,'pending_count':pending,'updated_at':now_iso()},eq={'id':jid,'user_id':user})

async def run_job(user,jid):
    job=db.one('jobs',user,eq={'id':jid});accounts={str(a['id']):a for a in db.rows('telegram_accounts',user,order=None)};stop=_stops[jid];pause=_pauses[jid];clients={};sources={}
    db.update('jobs',{'status':'RUNNING','updated_at':now_iso()},eq={'id':jid,'user_id':user});event(user,jid,'INFO','JOB','작업 시작')
    try:
        targets_all=db.rows('job_targets',user,eq={'job_id':jid},order='created_at');account_ids=sorted({str(x['assigned_account_id']) for x in targets_all if x.get('assigned_account_id')})
        for aid in account_ids:
            a=accounts.get(aid)
            if not a:continue
            c=await client_from_account(a);await c.connect();me=await c.get_me();clients[aid]=c
            bot,src=await prepare_source(c,job['bot_username'],dec(job['bot_token_enc']),me.id,job['message_text'],job['button_text'],job['button_url']);sources[aid]=(bot,src);event(user,jid,'INFO','SOURCE',f'텍스트+버튼 원본 준비 완료 message_id={src.id}',aid)
        while not stop.is_set():
            while pause.is_set() and not stop.is_set():await asyncio.sleep(.3)
            waiting=db.rows('job_targets',user,eq={'job_id':jid,'state':'WAITING'},order='created_at',limit=1)
            if not waiting:break
            t=waiting[0];tid=t['id'];aid=str(t['assigned_account_id']);c=clients.get(aid);bot,src=sources.get(aid,(None,None))
            db.update('job_targets',{'state':'PROCESSING','stage':'연락처 확인','updated_at':now_iso()},eq={'id':tid,'user_id':user});event(user,jid,'INFO','TARGET',f"{t['phone']} 연락처 확인",aid)
            try:
                uid=t.get('telegram_user_id')
                if not uid:
                    result=await c(functions.contacts.ImportContactsRequest([InputPhoneContact(client_id=random.randrange(1,2**63),phone=t['phone'],first_name=f'H24-{str(tid)[:8]}',last_name='')]))
                    if not result.users:raise RuntimeError('CONTACT_NOT_RESOLVED')
                    uid=int(result.users[0].id);db.update('contacts',{'telegram_user_id':uid,'assigned_account_id':aid,'state':'RESOLVED','detail':'Telegram UID 확인'},eq={'id':t['contact_id'],'user_id':user});db.update('job_targets',{'telegram_user_id':uid,'updated_at':now_iso()},eq={'id':tid,'user_id':user})
                if job.get('global_dedupe'):
                    history=db.rows('send_history',user,order='created_at',desc=True)
                    if any(h.get('phone')==t['phone'] or str(h.get('telegram_user_id'))==str(uid) for h in history):
                        db.update('job_targets',{'state':'SKIPPED','stage':'중복 제외','updated_at':now_iso()},eq={'id':tid,'user_id':user});event(user,jid,'INFO','DEDUPE',f"{t['phone']} 기존 성공 이력으로 제외",aid);recalc_job(user,jid);continue
                db.update('job_targets',{'stage':'발송 중','updated_at':now_iso()},eq={'id':tid,'user_id':user});peer=await c.get_input_entity(uid);sent=await c.forward_messages(peer,src.id,from_peer=bot);msg=sent[0] if isinstance(sent,list) else sent;mid=int(msg.id)
                db.update('job_targets',{'state':'SENT','stage':'완료','message_id':mid,'updated_at':now_iso()},eq={'id':tid,'user_id':user});db.insert('send_history',{'user_id':user,'phone':t['phone'],'telegram_user_id':uid,'account_id':aid,'job_id':jid,'message_id':mid});a=accounts.get(aid) or {};db.update('telegram_accounts',{'sent_count':int(a.get('sent_count') or 0)+1},eq={'id':aid,'user_id':user});a['sent_count']=int(a.get('sent_count') or 0)+1;event(user,jid,'INFO','SEND',f"{t['phone']} 발송 완료 message_id={mid}",aid);recalc_job(user,jid);await asyncio.sleep(random.uniform(float(job.get('delay_min') or 0),float(job.get('delay_max') or 0)))
            except Exception as e:
                if is_rate_error(e):
                    db.update('job_targets',{'state':'CHECK_REQUIRED','stage':'Telegram 제한','error_code':'TELEGRAM_RATE_LIMIT','error_detail':str(e),'updated_at':now_iso()},eq={'id':tid,'user_id':user});db.update('jobs',{'status':'PAUSED','stop_reason':'TELEGRAM_RATE_LIMIT','updated_at':now_iso()},eq={'id':jid,'user_id':user});event(user,jid,'CRITICAL','RATE_LIMIT',f'Telegram 제한 감지 / 전체 작업 중지 / {e}',aid);stop.set();break
                db.update('job_targets',{'state':'FAILED','stage':'실패','error_detail':str(e),'updated_at':now_iso()},eq={'id':tid,'user_id':user});event(user,jid,'ERROR','TARGET',f"{t['phone']} 실패 / {e}",aid);recalc_job(user,jid)
        current=db.one('jobs',user,eq={'id':jid})
        if current and current.get('status') not in ('PAUSED','STOPPED'):
            db.update('jobs',{'status':'COMPLETED','updated_at':now_iso()},eq={'id':jid,'user_id':user});event(user,jid,'INFO','JOB','작업 완료')
    finally:
        for c in clients.values():
            try:await c.disconnect()
            except Exception:pass
        _tasks.pop(jid,None);_stops.pop(jid,None);_pauses.pop(jid,None)

@app.post('/v1/jobs/{jid}/start')
async def start_job(jid:str,user=Depends(auth)):
    job=db.one('jobs',user,eq={'id':jid})
    if not job:raise HTTPException(404,'job not found')
    if jid in _tasks and not _tasks[jid].done():_pauses[jid].clear();db.update('jobs',{'status':'RUNNING','updated_at':now_iso()},eq={'id':jid,'user_id':user});return {'ok':True,'status':'RUNNING'}
    _stops[jid]=asyncio.Event();_pauses[jid]=asyncio.Event();_tasks[jid]=asyncio.create_task(run_job(user,jid));return {'ok':True,'status':'RUNNING'}

@app.post('/v1/jobs/{jid}/pause')
def pause_job(jid:str,user=Depends(auth)):
    if jid not in _tasks:raise HTTPException(409,'job not running')
    _pauses[jid].set();db.update('jobs',{'status':'PAUSED','updated_at':now_iso()},eq={'id':jid,'user_id':user});event(user,jid,'INFO','JOB','일시정지');return {'ok':True,'status':'PAUSED'}

@app.post('/v1/jobs/{jid}/stop')
def stop_job(jid:str,user=Depends(auth)):
    if jid in _stops:_stops[jid].set()
    db.update('jobs',{'status':'STOPPED','stop_reason':'USER_STOP','updated_at':now_iso()},eq={'id':jid,'user_id':user});event(user,jid,'INFO','JOB','사용자 중지');return {'ok':True,'status':'STOPPED'}
